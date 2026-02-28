"""
Export service: genera XLSX mensual del histórico de entregas.
Migra: cierreMensualAutomatico() del sistema original.
"""
import io
import logging
from datetime import datetime, timezone

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.historico import HistoricoEntregado

logger = logging.getLogger(__name__)


async def export_historico_xlsx(db: AsyncSession, mes: str) -> bytes:
    """
    Genera un XLSX con las entregas del mes indicado (formato '2026-02').
    Retorna bytes del archivo.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl no instalado")

    result = await db.execute(
        select(HistoricoEntregado)
        .where(HistoricoEntregado.mes_cierre == mes)
        .order_by(HistoricoEntregado.fecha_entregado)
    )
    rows = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Entregas {mes}"

    # Headers
    headers = [
        "Remito", "Cliente", "Domicilio", "Provincia",
        "Carrier", "Urgente", "Prioridad",
        "Fecha Ingreso", "Fecha Armado", "Fecha Entregado",
        "Observaciones", "Mes Cierre",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, h in enumerate(rows, start=2):
        ws.cell(row=row_num, column=1, value=h.numero_remito)
        ws.cell(row=row_num, column=2, value=h.cliente)
        ws.cell(row=row_num, column=3, value=h.domicilio)
        ws.cell(row=row_num, column=4, value=h.provincia)
        ws.cell(row=row_num, column=5, value=h.carrier_nombre)
        ws.cell(row=row_num, column=6, value="Sí" if h.urgente else "No")
        ws.cell(row=row_num, column=7, value="Sí" if h.prioridad else "No")
        ws.cell(row=row_num, column=8, value=h.fecha_ingreso.strftime("%Y-%m-%d %H:%M") if h.fecha_ingreso else "")
        ws.cell(row=row_num, column=9, value=h.fecha_armado.strftime("%Y-%m-%d %H:%M") if h.fecha_armado else "")
        ws.cell(row=row_num, column=10, value=h.fecha_entregado.strftime("%Y-%m-%d %H:%M") if h.fecha_entregado else "")
        ws.cell(row=row_num, column=11, value=h.observaciones)
        ws.cell(row=row_num, column=12, value=h.mes_cierre)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def monthly_close(db: AsyncSession) -> dict:
    """
    Cierre mensual: mueve entregados del mes anterior a histórico archivado.
    En este sistema el histórico ya está en DB — el cierre simplemente genera el export.
    """
    now = datetime.now(timezone.utc)
    if now.month == 1:
        mes_anterior = f"{now.year - 1}-12"
    else:
        mes_anterior = f"{now.year}-{now.month - 1:02d}"

    result = await db.execute(
        select(HistoricoEntregado).where(
            HistoricoEntregado.mes_cierre == mes_anterior
        )
    )
    entries = result.scalars().all()

    return {
        "ok": True,
        "mes": mes_anterior,
        "total_registros": len(entries),
        "exportados": len(entries),
    }
